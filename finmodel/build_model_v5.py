# -*- coding: utf-8 -*-
"""Финмодель v2 — Excel под движок тренажёра v2 (порт finmodel-trenazher.html).

Листы: «Сводка» (лестница 4 пресетов, живые ссылки) + 4 листа-пресета.
Каждый лист-пресет — живая модель: параметры сверху (жёлтое = главные рычаги,
синее = остальные, всё можно править), таблица 41 месяц (авг-26 → дек-29)
живыми формулами, KPI-плитки как в тренажёре.

Генератор трёхступенчато сверяет сам себя, файл пишется только если всё сошлось:
  1) python-порт движка == опубликованные цифры пресетов тренажёра;
  2) каждая формула Excel вычисляется встроенным мини-вычислителем;
  3) вычисленные формулы == python-порт (до копейки).
"""
import math
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

OUT = '/Users/nikolaymacbook/Claude/mama-ryadom-bot/temp/Финмодель v2.xlsx'

# ============================ 1. ДВИЖОК (порт тренажёра) ============================

SEAS = {1: (1.0, 0.6), 2: (1.4, 1.0), 3: (1.1, 1.0), 4: (1.0, 1.0),
        5: (1.0, 1.0), 6: (0.6, 0.9), 7: (0.5, 0.9), 8: (0.7, 1.0),
        9: (1.2, 1.0), 10: (1.1, 1.0), 11: (1.0, 1.0), 12: (0.8, 0.8)}
N = 41
MONTHS = []
_y, _m = 2026, 8
for _ in range(N):
    MONTHS.append((_y, _m))
    _m += 1
    if _m == 13:
        _m, _y = 1, _y + 1
RU = {1: 'янв', 2: 'фев', 3: 'мар', 4: 'апр', 5: 'май', 6: 'июн',
      7: 'июл', 8: 'авг', 9: 'сен', 10: 'окт', 11: 'ноя', 12: 'дек'}
MDATES = [datetime(y, m, 1) for y, m in MONTHS]


def mlabel(i):
    y, m = MONTHS[i]
    return f"{RU[m]}-{y % 100:02d}"


BASE = dict(
    cpl=800, capLeads=130, cp=0.13, trial=0.50, pay=0.93,
    org0=23, orgg=0.0, orgCap=25, co=0.23, ch2=0, cpl2=1000,
    boostLeads=220, boostCpl=1200, boostCh2=80,
    bump1=0.0, bump2=0.0, trialPrice=900,
    other=435000, fot=350000, rateG=500, rateC=100, gmin=30, fill=4.0, visits=4.33,
    taxAug=38260, taxDec=40000,
    plan=3000000, spent=1551733, rebuild=1700000, startBase=0, cap=250,
    life=3.6, nomin=5200, ref=0.0, boostM=0, usn=1)
DOKR = {**BASE, **dict(life=5.4, nomin=6150, ref=0.06, org0=15, orgg=0.06,
                       orgCap=60, ch2=40, bump1=0.08, bump2=0.06, fot=300000)}
FORS = {**DOKR, **dict(boostM=18)}
DREAM = {**FORS, **dict(life=6.0)}

PRESETS = [
    ('Как сейчас', BASE,
     'жизнь 3,6 мес · 5 200 ₽/мес · рефералки нет · форсажа нет · органика 23 замороженная · ФОТ 350'),
    ('Докрученная', DOKR,
     'жизнь 5,4 · 6 150 ₽ · реф 6% · органика 15→60 (+6%/мес) · 2-й канал 40 · индексации 8/6% · ФОТ 300'),
    ('+ Форсаж', FORS,
     'докрученная + форсаж 18 мес: таргет 220 @ 1 200 ₽ + 2-й канал 80'),
    ('Мечта — жизнь 6', DREAM,
     'то же + удержание получилось: жизнь 6 мес'),
]


def simulate(s):
    """Точный порт simulate() из finmodel-trenazher.html (проценты — долями)."""
    A, cash = float(s['startBase']), float(s['plan'] - s['spent'])
    usn = s['usn']
    out = dict(A=[], rev=[], pnl=[], cash=[])
    for i in range(N):
        m = MONTHS[i][1]
        si, sc = SEAS[m]
        pf = ((1 + s['bump1']) if i >= 13 else 1.0) * ((1 + s['bump2']) if i >= 25 else 1.0)
        boost = i < s['boostM']
        tgt = s['boostLeads'] if boost else s['capLeads']
        cpl = s['boostCpl'] if boost else s['cpl']
        ch2n = (s['boostCh2'] if boost else s['ch2']) if i >= 2 else 0
        cpl2 = s['boostCpl'] if boost else s['cpl2']
        lp = tgt * si
        lo = min(s['orgCap'], s['org0'] * (1 + s['orgg']) ** i) * si
        lc = ch2n * si
        nw = ((lp + lc) * s['cp'] + lo * s['co']) * sc * s['pay'] + s['ref'] * A * sc
        A = min(s['cap'], A * (1 - 1 / s['life']) + nw)
        tr = (lp + lo + lc) * s['trial'] * sc
        rev = A * s['nomin'] * pf + max(0.0, tr - nw) * s['trialPrice']
        teach = max(s['gmin'], A / s['fill']) * s['visits'] * s['rateG'] + A * s['visits'] * s['rateC']
        tax = (s['taxAug'] if i == 0 else 0) + (s['taxDec'] if i == 4 else 0) + (0.06 * rev if usn else 0)
        mkt = tgt * cpl + ch2n * cpl2
        pnl = rev - mkt - s['other'] - s['fot'] - teach - tax
        cash += pnl - (s['rebuild'] if i == 0 else 0)
        out['A'].append(A); out['rev'].append(rev); out['pnl'].append(pnl); out['cash'].append(cash)
    return out


def breakeven_iter(s):
    """Порт breakeven() тренажёра: первый целый A, где выручка покрывает расходы."""
    bud = s['capLeads'] * s['cpl'] + s['ch2'] * s['cpl2']
    for A in range(0, int(s['cap']) + 1):
        rev = A * s['nomin']
        costs = (bud + s['other'] + s['fot']
                 + max(s['gmin'], A / s['fill']) * s['visits'] * s['rateG']
                 + A * s['visits'] * s['rateC'] + (0.06 * rev if s['usn'] else 0))
        if rev >= costs:
            return A
    return None


def breakeven_raw(s):
    """Замкнутая форма той же ТБ (её и пишем формулой в Excel). 1e9 = недостижимо."""
    bud = s['capLeads'] * s['cpl'] + s['ch2'] * s['cpl2']
    coef1 = s['nomin'] * (1 - 0.06 * s['usn']) - s['visits'] * s['rateC']
    if coef1 <= 0:
        return 1e9
    a1 = (bud + s['other'] + s['fot'] + s['gmin'] * s['visits'] * s['rateG']) / coef1
    if a1 <= s['gmin'] * s['fill']:
        return a1
    coef2 = coef1 - s['visits'] * s['rateG'] / s['fill']
    if coef2 <= 0:
        return 1e9
    return (bud + s['other'] + s['fot']) / coef2


def kpis(s):
    r = simulate(s)
    fp = next((i for i, p in enumerate(r['pnl']) if p > 0), None)
    f5 = next((i for i, p in enumerate(r['pnl']) if p >= 500000), None)
    dno = min(r['cash'])
    dno_i = r['cash'].index(dno)
    shortfall = max(0.0, -dno)
    dipped = any(c < 0 for c in r['cash'])
    recov = None
    if dipped:
        seen_neg = False
        for i, c in enumerate(r['cash']):
            if c < 0:
                seen_neg = True
            elif seen_neg:
                recov = i
                break
    idx = lambda y, m: (y - 2026) * 12 + (m - 8)
    return dict(r=r, fp=fp, f5=f5, dno=dno, dno_i=dno_i, shortfall=shortfall,
                total=s['plan'] + shortfall, dipped=dipped, recov=recov,
                cli27=r['A'][idx(2027, 12)], cli28=r['A'][idx(2028, 12)], cli29=r['A'][idx(2029, 12)],
                y28=sum(r['pnl'][idx(2028, 1):idx(2029, 1)]), y29=sum(r['pnl'][idx(2029, 1):]),
                cash_end=r['cash'][-1], be=breakeven_iter(s), be_raw=breakeven_raw(s))


# --- сверка порта с опубликованными цифрами тренажёра (память проекта, УСН вкл) ---

def check_port():
    exp = {  # имя: (первый плюс, первые 500к, дно млн, дно-месяц, всего млн, касса дек-29 млн)
        'Как сейчас':      (None,     None,     None, None,     None, -26.3),
        'Докрученная':     ('окт-27', 'апр-29', -5.3, 'сен-27',  8.3,   2.7),
        '+ Форсаж':        ('апр-27', 'фев-28', -3.9, None,      6.9,   8.2),
        'Мечта — жизнь 6': ('мар-27', None,     -3.7, None,      6.7,  10.1),
    }
    for name, s, _ in PRESETS:
        k = kpis(s)
        e_fp, e_f5, e_dno, e_dno_m, e_tot, e_cash = exp[name]
        got_fp = mlabel(k['fp']) if k['fp'] is not None else None
        got_f5 = mlabel(k['f5']) if k['f5'] is not None else None
        def close(a, b, tol=0.07):
            return b is None or abs(a - b) <= tol
        ok = ((e_fp is None or got_fp == e_fp) and (e_f5 is None or got_f5 == e_f5)
              and close(k['dno'] / 1e6, e_dno) and (e_dno_m is None or mlabel(k['dno_i']) == e_dno_m)
              and close(k['total'] / 1e6, e_tot) and close(k['cash_end'] / 1e6, e_cash))
        print(f"  порт {name:18s} плюс {got_fp or '—':7s} 500к {got_f5 or '—':7s} "
              f"дно {k['dno']/1e6:+.1f} всего {k['total']/1e6:.1f} касса29 {k['cash_end']/1e6:+.1f} "
              f"ТБ {k['be']} {'✓' if ok else '✗ РАСХОЖДЕНИЕ С ТРЕНАЖЁРОМ'}")
        assert ok, f"порт разошёлся с тренажёром: {name}"
        raw = k['be_raw']
        be_closed = None if raw > s['cap'] else math.ceil(raw - 1e-12)
        assert be_closed == k['be'], f"замкнутая ТБ != переборной: {name} {be_closed} {k['be']}"


# ============================ 2. МИНИ-DSL ФОРМУЛ ============================
# Каждая ячейка = выражение: рендерится в текст формулы И вычисляется питоном.

class XlErr(Exception):
    pass


def wrap(x):
    if isinstance(x, Node):
        return x
    if isinstance(x, (int, float)):
        return Num(x)
    if isinstance(x, str):
        return Txt(x)
    if isinstance(x, datetime):
        return Lit(x)
    raise TypeError(repr(x))


class Node:
    prec = 9

    def __add__(self, o): return Bin('+', self, wrap(o))
    def __radd__(self, o): return Bin('+', wrap(o), self)
    def __sub__(self, o): return Bin('-', self, wrap(o))
    def __rsub__(self, o): return Bin('-', wrap(o), self)
    def __mul__(self, o): return Bin('*', self, wrap(o))
    def __rmul__(self, o): return Bin('*', wrap(o), self)
    def __truediv__(self, o): return Bin('/', self, wrap(o))
    def __rtruediv__(self, o): return Bin('/', wrap(o), self)
    def __pow__(self, o): return Bin('^', self, wrap(o))
    def __neg__(self): return Neg(self)

    def eq(self, o): return Bin('=', self, wrap(o))
    def le(self, o): return Bin('<=', self, wrap(o))
    def ge(self, o): return Bin('>=', self, wrap(o))
    def lt(self, o): return Bin('<', self, wrap(o))
    def gt(self, o): return Bin('>', self, wrap(o))


class Num(Node):
    def __init__(self, v): self.v = v
    def render(self, sheet):
        if isinstance(self.v, int) or float(self.v).is_integer():
            return str(int(self.v))
        return repr(float(self.v))
    def ev(self, ctx, sheet): return float(self.v)


class Txt(Node):
    def __init__(self, v): self.v = v
    def render(self, sheet): return '"' + self.v + '"'
    def ev(self, ctx, sheet): return self.v


class Lit(Node):  # значение-литерал ячейки (дата и т.п.), в формулы не рендерится
    def __init__(self, v): self.v = v
    def render(self, sheet): raise RuntimeError('литерал не рендерится в формулу')
    def ev(self, ctx, sheet): return self.v


PREC = {'^': 4, '*': 3, '/': 3, '+': 2, '-': 2, '=': 1, '<': 1, '>': 1, '<=': 1, '>=': 1}


class Bin(Node):
    def __init__(self, op, l, r):
        self.op, self.l, self.r = op, l, r
        self.prec = PREC[op]

    def render(self, sheet):
        ls = self.l.render(sheet)
        rs = self.r.render(sheet)
        if self.l.prec < self.prec:
            ls = '(' + ls + ')'
        if self.r.prec < self.prec or (self.r.prec == self.prec and self.op in ('-', '/', '^')):
            rs = '(' + rs + ')'
        return ls + self.op + rs

    def ev(self, ctx, sheet):
        a, b = self.l.ev(ctx, sheet), self.r.ev(ctx, sheet)
        if self.op == '+': return a + b
        if self.op == '-': return a - b
        if self.op == '*': return a * b
        if self.op == '/':
            if b == 0:
                raise XlErr('#DIV/0!')
            return a / b
        if self.op == '^': return a ** b
        if self.op == '=': return a == b
        if self.op == '<': return a < b
        if self.op == '>': return a > b
        if self.op == '<=': return a <= b
        if self.op == '>=': return a >= b


class Neg(Node):
    prec = 3.5
    def __init__(self, x): self.x = x
    def render(self, sheet):
        xs = self.x.render(sheet)
        if self.x.prec < self.prec:
            xs = '(' + xs + ')'
        return '-' + xs
    def ev(self, ctx, sheet): return -self.x.ev(ctx, sheet)


def q(sheetname):
    return "'" + sheetname + "'!"


class Ref(Node):
    def __init__(self, col, row, sheet=None, absolute=True):
        self.col, self.row, self.sheet, self.absolute = col, row, sheet, absolute

    def render(self, sheet):
        s = q(self.sheet) if (self.sheet and self.sheet != sheet) else ''
        d = '$' if self.absolute else ''
        return f"{s}{d}{self.col}{d}{self.row}"

    def ev(self, ctx, sheet):
        return ctx.value(self.sheet or sheet, self.col, self.row)


class Rng(Node):
    def __init__(self, col1, row1, col2, row2, sheet=None):
        assert col1 == col2, 'диапазоны только по одному столбцу'
        self.col, self.r1, self.r2, self.sheet = col1, row1, row2, sheet

    def render(self, sheet):
        s = q(self.sheet) if (self.sheet and self.sheet != sheet) else ''
        return f"{s}${self.col}${self.r1}:{s}${self.col}${self.r2}" if not s else \
               f"{s}${self.col}${self.r1}:${self.col}${self.r2}"

    def ev(self, ctx, sheet):
        sh = self.sheet or sheet
        return [ctx.value(sh, self.col, r) for r in range(self.r1, self.r2 + 1)]


class Fn(Node):
    def __init__(self, name, *args):
        self.name = name
        self.args = [wrap(a) for a in args]

    def render(self, sheet):
        return self.name + '(' + ','.join(a.render(sheet) for a in self.args) + ')'

    def ev(self, ctx, sheet):
        nm = self.name
        if nm == 'IF':
            c = self.args[0].ev(ctx, sheet)
            return self.args[1].ev(ctx, sheet) if c else self.args[2].ev(ctx, sheet)
        if nm == 'IFERROR':
            try:
                return self.args[0].ev(ctx, sheet)
            except XlErr:
                return self.args[1].ev(ctx, sheet)
        a = [x.ev(ctx, sheet) for x in self.args]
        flat = []
        for v in a:
            flat.extend(v if isinstance(v, list) else [v])
        if nm == 'MIN': return min(flat)
        if nm == 'MAX': return max(flat)
        if nm == 'SUM': return sum(flat)
        if nm == 'ABS': return abs(a[0])
        if nm == 'AND': return all(a)
        if nm == 'ROUND':
            x = a[0]
            return math.floor(x + 0.5) if x >= 0 else -math.floor(-x + 0.5)
        if nm == 'ROUNDUP':
            x = a[0]
            return math.ceil(x) if x >= 0 else -math.ceil(-x)
        if nm == 'MATCH':
            target, rng = a[0], a[1]
            for i, v in enumerate(rng):
                if v == target:
                    return i + 1
            raise XlErr('#N/A')
        if nm == 'INDEX':
            return a[0][int(a[1]) - 1]
        if nm == 'COUNTIF':
            rng, crit = a[0], a[1]
            op = crit[0]
            thr = float(crit[1:])
            if op == '<':
                return sum(1 for v in rng if isinstance(v, (int, float)) and v < thr)
            if op == '>':
                return sum(1 for v in rng if isinstance(v, (int, float)) and v > thr)
            raise RuntimeError('COUNTIF: ' + crit)
        raise RuntimeError('нет функции ' + nm)


IF, MINF, MAXF, SUMF = (lambda *a: Fn('IF', *a)), (lambda *a: Fn('MIN', *a)), \
                       (lambda *a: Fn('MAX', *a)), (lambda *a: Fn('SUM', *a))


class Registry:
    """Все ячейки книги: (лист, колонка, строка) -> Node. Вычисление с мемо."""

    def __init__(self):
        self.cells = {}
        self.memo = {}
        self.busy = set()

    def put(self, sheet, col, row, node):
        self.cells[(sheet, col, row)] = node

    def value(self, sheet, col, row):
        key = (sheet, col, row)
        if key in self.memo:
            return self.memo[key]
        if key not in self.cells:
            raise RuntimeError(f'пустая ячейка в формуле: {key}')
        if key in self.busy:
            raise RuntimeError(f'циклическая ссылка: {key}')
        self.busy.add(key)
        v = self.cells[key].ev(self, sheet)
        self.busy.discard(key)
        self.memo[key] = v
        return v


REG = Registry()

# ============================ 3. СТИЛИ ============================

MAIN = PatternFill('solid', fgColor='FFD966')   # главные рычаги
BLUE = PatternFill('solid', fgColor='DDEBF7')   # параметры (можно править)
GREY = PatternFill('solid', fgColor='F2F2F2')   # заголовки
RED = PatternFill('solid', fgColor='F8CBAD')    # касса < 0
HDR = Font(bold=True)
H1 = Font(bold=True, size=14)
SMALL = Font(size=9, color='808080')
MONEY = '#,##0;[Red]-#,##0'
DATEF = 'MMM YY'
thin = Side(style='thin', color='CCCCCC')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def put(ws, col, row, value, fmt=None, fill=None, font=None, border=False, wrap_text=False):
    """Пишет в лист и в реестр. value: Node (формула) или литерал."""
    c = ws.cell(row=row, column=openpyxl.utils.column_index_from_string(col))
    if isinstance(value, Node):
        c.value = '=' + value.render(ws.title)
        REG.put(ws.title, col, row, value)
    else:
        c.value = value
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            REG.put(ws.title, col, row, Num(value))
        elif isinstance(value, datetime):
            REG.put(ws.title, col, row, Lit(value))
        elif isinstance(value, str):
            REG.put(ws.title, col, row, Txt(value))
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if font:
        c.font = font
    if border:
        c.border = BORDER
    if wrap_text:
        c.alignment = Alignment(wrap_text=True, vertical='center')
    return c


# ============================ 4. ЛИСТ-ПРЕСЕТ ============================

TR0 = 30            # первая строка таблицы месяцев
TRE = TR0 + N - 1   # последняя (=70)
ROW_DEC27, ROW_DEC28, ROW_DEC29 = TR0 + 16, TR0 + 28, TR0 + 40
Y26 = (TR0, TR0 + 4)
Y27 = (TR0 + 5, TR0 + 16)
Y28 = (TR0 + 17, TR0 + 28)
Y29 = (TR0 + 29, TR0 + 40)

# параметры: key -> (label, unit, fmt, side, main)
PARAM_LAYOUT = [
    ('L', 'ГЛАВНЫЕ РЫЧАГИ (как в тренажёре)', None),
    ('life', 'Сколько месяцев ходит средний ребёнок', 'мес', '0.0', True),
    ('nomin', '₽ с клиента в месяц (тарифный микс)', '₽/мес', MONEY, True),
    ('ref', 'Рефералка: % базы приводит друга в месяц', '%/мес', '0%', True),
    ('boostM', 'Форсаж: месяцев с открытия (0 = выкл)', 'мес', '0', True),
    ('usn', 'УСН 6% с выручки (1 = вкл, честный режим)', '1/0', '0', True),
    ('L', '', None),
    ('L', 'Базовый маркетинг', None),
    ('cpl', 'CPL — цена сырого лида (таргет)', '₽', MONEY, False),
    ('capLeads', 'Потолок платных лидов (проверенный ~130)', 'шт/мес', '0', False),
    ('cp', 'Конверсия платного лида в покупку', '', '0%', False),
    ('co', 'Конверсия органической заявки', '', '0%', False),
    ('trial', 'Доля лидов → пробное', '', '0%', False),
    ('pay', 'Коэфф. «зачислен → оплатил»', '', '0.00', False),
    ('L', '', None),
    ('L', 'Органика и второй канал', None),
    ('org0', 'Органические заявки на старте', 'шт/мес', '0', False),
    ('orgg', 'Рост органики', '%/мес', '0%', False),
    ('orgCap', 'Потолок органики', 'шт/мес', '0', False),
    ('ch2', 'Второй платный канал (заявок, с окт-26)', 'шт/мес', '0', False),
    ('cpl2', 'CPL второго канала', '₽', MONEY, False),
    ('L', '', None),
    ('L', 'Форсаж — параметры', None),
    ('boostLeads', 'Таргет в форсаже', 'шт/мес', '0', False),
    ('boostCpl', 'CPL в форсаже (оба канала)', '₽', MONEY, False),
    ('boostCh2', 'Второй канал в форсаже', 'шт/мес', '0', False),
]
PARAM_LAYOUT_R = [
    ('R', 'Цены', None),
    ('bump1', 'Индексация прайса с сен-2027', '', '0%', False),
    ('bump2', 'Ещё индексация с сен-2028', '', '0%', False),
    ('trialPrice', 'Цена пробного (некупившие платят)', '₽', MONEY, False),
    ('R', '', None),
    ('R', 'Расходы', None),
    ('other', 'Прочие постоянные (аренда и пр.)', '₽/мес', MONEY, False),
    ('fot', 'ФОТ фикс-ядро', '₽/мес', MONEY, False),
    ('rateG', 'Ставка тренера за группу (и пустую)', '₽', MONEY, False),
    ('rateC', 'Ставка тренера за ребёнка', '₽', MONEY, False),
    ('gmin', 'Мин. групп в неделю (стартовая сетка)', 'шт', '0', False),
    ('fill', 'Средняя наполняемость группы', 'чел', '0.0', False),
    ('visits', 'Посещений клиента в месяц', 'шт', '0.00', False),
    ('taxAug', 'Налог исходника: авг-26 (взносы ИП)', '₽', MONEY, False),
    ('taxDec', 'Налог исходника: дек-26', '₽', MONEY, False),
    ('R', '', None),
    ('R', 'Старт и деньги', None),
    ('plan', 'Финансирование всего', '₽', MONEY, False),
    ('spent', 'Потрачено до открытия (факт 03.07.26)', '₽', MONEY, False),
    ('rebuild', 'Достройка до открытия (август)', '₽', MONEY, False),
    ('startBase', 'Стартовая база клиентов', 'чел', '0', False),
    ('cap', 'Вместимость (действующих абонементов)', 'чел', '0', False),
]

COLS = [  # (col, заголовок, формат)
    ('A', '№', '0'),
    ('B', 'Месяц', DATEF),
    ('C', 'Форсаж (1/0)', '0'),
    ('D', 'Сезон: приток', '0.0'),
    ('E', 'Сезон: конверсия', '0.0'),
    ('F', 'Индекс прайса', '0.00'),
    ('G', 'Лиды: таргет', '0'),
    ('H', 'Лиды: 2-й канал', '0'),
    ('I', 'Заявки: органика', '0'),
    ('J', 'Лидов всего', '0'),
    ('K', 'Пробных занятий', '0'),
    ('L', 'НОВЫХ клиентов (купили)', '0.0'),
    ('M', '— из них рефералы', '0.0'),
    ('N', 'ДЕЙСТВУЮЩИХ клиентов', '0'),
    ('O', 'Выручка: абонементы', MONEY),
    ('P', 'Выручка: пробные', MONEY),
    ('Q', 'ИТОГО выручка', MONEY),
    ('R', 'Маркетинг', MONEY),
    ('S', 'Тренеры (500/гр + 100/реб)', MONEY),
    ('T', 'ФОТ фикс-ядро', MONEY),
    ('U', 'Прочие постоянные', MONEY),
    ('V', 'Налоги (фикс + УСН)', MONEY),
    ('W', 'Расходы всего', MONEY),
    ('X', 'РЕЗУЛЬТАТ месяца', MONEY),
    ('Y', 'КАССА на конец месяца', MONEY),
    ('AA', 'флаг: плюс', '0'),
    ('AB', 'флаг: ≥500к', '0'),
    ('AC', 'флаг: касса снова ≥0', '0'),
]

KPI_ROWS = [  # (row, label, формат)
    (4, 'Первый месяц в плюсе', DATEF),
    (5, 'Первые 500 тыс./мес', DATEF),
    (6, 'Дно кассы', MONEY),
    (7, '…когда дно', DATEF),
    (8, 'Не хватает денег (сверх финансирования)', MONEY),
    (9, 'ВСЕГО в проект (финансирование + дыра)', MONEY),
    (10, 'Клиентов к дек-2027', '0'),
    (11, 'Клиентов к дек-2028', '0'),
    (12, 'Клиентов к дек-2029', '0'),
    (13, 'Прибыль за 2028', MONEY),
    (14, 'Прибыль за 2029', MONEY),
    (15, 'КАССА на дек-2029', MONEY),
    (16, 'Касса снова ≥ 0', DATEF),
    (17, 'Точка безубыточности, клиентов', '0'),
    (18, 'Новых платящих/мес для нуля (~ТБ/жизнь)', '0'),
    (19, 'С клиента за жизнь', MONEY),
]


def build_preset_sheet(wb, name, preset, desc):
    ws = wb.create_sheet(name)
    ws.sheet_view.zoomScale = 90
    widths = {'A': 4.5, 'B': 9, 'C': 7.5, 'D': 8, 'E': 9, 'F': 8, 'G': 8, 'H': 8, 'I': 9,
              'J': 8, 'K': 9, 'L': 10, 'M': 10, 'N': 11, 'O': 12, 'P': 10, 'Q': 12,
              'R': 11, 'S': 11, 'T': 11, 'U': 11, 'V': 11, 'W': 12, 'X': 12, 'Y': 13,
              'Z': 2, 'AA': 6, 'AB': 6, 'AC': 6}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.column_dimensions['B'].width = 9

    put(ws, 'B', 1, f'{name} — {desc}', font=H1)
    put(ws, 'B', 2, 'Жёлтое — главные рычаги, синее — остальные параметры (можно править). '
                    'Всё прочее — живые формулы. Движок и цифры = тренажёр v2.', font=SMALL)

    # --- параметры (лево B/C/D и право F/G/H) ---
    P = {}

    def emit_params(layout, lc, vc, uc, start_row):
        r = start_row
        for item in layout:
            if item[0] in ('L', 'R'):
                _, title, _ = item
                if title:
                    put(ws, lc, r, title, fill=GREY, font=HDR)
                r += 1
                continue
            key, label, unit, fmt, main = item
            put(ws, lc, r, label)
            put(ws, vc, r, preset[key], fmt=fmt, border=True,
                fill=MAIN if main else BLUE, font=HDR if main else None)
            put(ws, uc, r, unit, font=SMALL)
            P[key] = Ref(vc, r)
            r += 1
        return r

    emit_params(PARAM_LAYOUT, 'B', 'C', 'D', 3)
    emit_params(PARAM_LAYOUT_R, 'F', 'G', 'H', 3)

    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['F'].width = 34
    for cdim, w in (('C', 10), ('D', 7), ('G', 10), ('H', 7)):
        ws.column_dimensions[cdim].width = w

    # --- таблица месяцев ---
    put(ws, 'A', TR0 - 2, 'МОДЕЛЬ ПО МЕСЯЦАМ (все строки — формулы; движок тренажёра v2)', font=HDR, fill=GREY)
    for col, title, fmt in COLS:
        c = put(ws, col, TR0 - 1, title, fill=GREY, font=Font(bold=True, size=9), border=True, wrap_text=True)
        c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    ws.row_dimensions[TR0 - 1].height = 40
    put(ws, 'AA', TR0 - 2, 'служебные →', font=SMALL)

    for i in range(N):
        r = TR0 + i
        no = Ref('A', r, absolute=False)
        m = MONTHS[i][1]
        si, sc = SEAS[m]
        prevN = P['startBase'] if i == 0 else Ref('N', r - 1, absolute=False)

        put(ws, 'A', r, i + 1, fmt='0', border=True)
        put(ws, 'B', r, MDATES[i], fmt=DATEF, border=True, font=HDR)
        put(ws, 'C', r, IF(no.le(P['boostM']), 1, 0), fmt='0', border=True)
        put(ws, 'D', r, si, fmt='0.0', border=True, fill=BLUE)
        put(ws, 'E', r, sc, fmt='0.0', border=True, fill=BLUE)
        F_ = IF(no.ge(14), 1 + P['bump1'], 1) * IF(no.ge(26), 1 + P['bump2'], 1)
        put(ws, 'F', r, F_, fmt='0.00', border=True)

        C_, D_, E_ = Ref('C', r, absolute=False), Ref('D', r, absolute=False), Ref('E', r, absolute=False)
        put(ws, 'G', r, IF(C_.eq(1), P['boostLeads'], P['capLeads']) * D_, fmt='0', border=True)
        put(ws, 'H', r, IF(no.ge(3), IF(C_.eq(1), P['boostCh2'], P['ch2']), 0) * D_, fmt='0', border=True)
        put(ws, 'I', r, MINF(P['orgCap'], P['org0'] * (1 + P['orgg']) ** (no - 1)) * D_, fmt='0', border=True)
        G_, H_, I_ = (Ref(c_, r, absolute=False) for c_ in 'GHI')
        put(ws, 'J', r, G_ + H_ + I_, fmt='0', border=True)
        put(ws, 'K', r, Ref('J', r, absolute=False) * P['trial'] * E_, fmt='0', border=True)
        put(ws, 'M', r, P['ref'] * prevN * E_, fmt='0.0', border=True)
        put(ws, 'L', r, ((G_ + H_) * P['cp'] + I_ * P['co']) * E_ * P['pay'] + Ref('M', r, absolute=False),
            fmt='0.0', border=True)
        L_ = Ref('L', r, absolute=False)
        put(ws, 'N', r, MINF(P['cap'], prevN * (1 - 1 / P['life']) + L_), fmt='0', border=True, font=HDR)
        N_ = Ref('N', r, absolute=False)
        put(ws, 'O', r, N_ * P['nomin'] * Ref('F', r, absolute=False), fmt=MONEY, border=True)
        put(ws, 'P', r, MAXF(0, Ref('K', r, absolute=False) - L_) * P['trialPrice'], fmt=MONEY, border=True)
        Q_ = Ref('O', r, absolute=False) + Ref('P', r, absolute=False)
        put(ws, 'Q', r, Q_, fmt=MONEY, border=True, font=HDR)
        put(ws, 'R', r, IF(C_.eq(1), P['boostLeads'] * P['boostCpl'], P['capLeads'] * P['cpl'])
            + IF(no.ge(3), 1, 0) * IF(C_.eq(1), P['boostCh2'] * P['boostCpl'], P['ch2'] * P['cpl2']),
            fmt=MONEY, border=True)
        put(ws, 'S', r, MAXF(P['gmin'], N_ / P['fill']) * P['visits'] * P['rateG'] + N_ * P['visits'] * P['rateC'],
            fmt=MONEY, border=True)
        put(ws, 'T', r, P['fot'] + Num(0), fmt=MONEY, border=True)
        put(ws, 'U', r, P['other'] + Num(0), fmt=MONEY, border=True)
        put(ws, 'V', r, IF(no.eq(1), P['taxAug'], 0) + IF(no.eq(5), P['taxDec'], 0)
            + IF(P['usn'].eq(1), Num(0.06) * Ref('Q', r, absolute=False), 0), fmt=MONEY, border=True)
        W_ = Ref('R', r, absolute=False) + Ref('S', r, absolute=False) + Ref('T', r, absolute=False) \
            + Ref('U', r, absolute=False) + Ref('V', r, absolute=False)
        put(ws, 'W', r, W_, fmt=MONEY, border=True)
        X_ = Ref('Q', r, absolute=False) - Ref('W', r, absolute=False)
        put(ws, 'X', r, X_, fmt=MONEY, border=True, font=HDR)
        if i == 0:
            cash = P['plan'] - P['spent'] + Ref('X', r, absolute=False) - P['rebuild']
        else:
            cash = Ref('Y', r - 1, absolute=False) + Ref('X', r, absolute=False)
        put(ws, 'Y', r, cash, fmt=MONEY, border=True, font=HDR)

        Xr, Yr = Ref('X', r, absolute=False), Ref('Y', r, absolute=False)
        put(ws, 'AA', r, IF(Xr.gt(0), 1, 0), fmt='0', font=SMALL)
        put(ws, 'AB', r, IF(Xr.ge(500000), 1, 0), fmt='0', font=SMALL)
        if i == 0:
            put(ws, 'AC', r, 0, fmt='0', font=SMALL)
        else:
            put(ws, 'AC', r, IF(Fn('AND', Yr.ge(0), Fn('COUNTIF', Rng('Y', TR0, 'Y', r - 1), '<0').gt(0)), 1, 0),
                fmt='0', font=SMALL)

    ws.conditional_formatting.add(
        f'Y{TR0}:Y{TRE}', CellIsRule(operator='lessThan', formula=['0'], fill=RED))
    ws.freeze_panes = 'C1'

    # --- KPI-плитки ---
    months_rng = Rng('B', TR0, 'B', TRE)
    cash_rng = Rng('Y', TR0, 'Y', TRE)
    put(ws, 'J', 3, 'ИТОГИ (плитки тренажёра)', fill=GREY, font=HDR)
    put(ws, 'K', 3, '', fill=GREY)
    kpi = {
        4: Fn('IFERROR', Fn('INDEX', months_rng, Fn('MATCH', 1, Rng('AA', TR0, 'AA', TRE), 0)), 'нет до 2030'),
        5: Fn('IFERROR', Fn('INDEX', months_rng, Fn('MATCH', 1, Rng('AB', TR0, 'AB', TRE), 0)), '—'),
        6: MINF(cash_rng),
        7: Fn('INDEX', months_rng, Fn('MATCH', MINF(cash_rng), cash_rng, 0)),
        8: MAXF(0, -MINF(cash_rng)),
        9: P['plan'] + Ref('K', 8),
        10: Ref('N', ROW_DEC27) + Num(0),
        11: Ref('N', ROW_DEC28) + Num(0),
        12: Ref('N', ROW_DEC29) + Num(0),
        13: SUMF(Rng('X', Y28[0], 'X', Y28[1])),
        14: SUMF(Rng('X', Y29[0], 'X', Y29[1])),
        15: Ref('Y', TRE) + Num(0),
        16: IF(MINF(cash_rng).ge(0), 'не проседает',
               Fn('IFERROR', Fn('INDEX', months_rng, Fn('MATCH', 1, Rng('AC', TR0, 'AC', TRE), 0)), 'после 2029')),
        17: IF(Ref('M', 8).gt(P['cap']), 'недостиж.', Fn('ROUNDUP', Ref('M', 8), 0)),
        18: IF(Ref('M', 8).gt(P['cap']), '—', Fn('ROUND', Ref('M', 8) / P['life'], 0)),
        19: P['nomin'] * P['life'],
    }
    for row, label, fmt in KPI_ROWS:
        put(ws, 'J', row, label)
        put(ws, 'K', row, kpi[row], fmt=fmt, font=HDR, border=True)

    # служебный расчёт ТБ (замкнутая форма перебора из тренажёра)
    put(ws, 'L', 3, '(служебный расчёт ТБ — не трогать)', font=SMALL)
    bud = P['capLeads'] * P['cpl'] + P['ch2'] * P['cpl2']
    put(ws, 'M', 4, P['nomin'] * (1 - Num(0.06) * P['usn']) - P['visits'] * P['rateC'], fmt='0.00', font=SMALL)
    put(ws, 'M', 5, IF(Ref('M', 4).le(0), 1000000000,
                       (bud + P['other'] + P['fot'] + P['gmin'] * P['visits'] * P['rateG']) / Ref('M', 4)),
        fmt='0.0', font=SMALL)
    put(ws, 'M', 6, Ref('M', 4) - P['visits'] * P['rateG'] / P['fill'], fmt='0.00', font=SMALL)
    put(ws, 'M', 7, IF(Ref('M', 6).le(0), 1000000000, (bud + P['other'] + P['fot']) / Ref('M', 6)),
        fmt='0.0', font=SMALL)
    put(ws, 'M', 8, IF(Ref('M', 5).le(P['gmin'] * P['fill']), Ref('M', 5), Ref('M', 7)), fmt='0.0', font=SMALL)

    # --- годовые итоги ---
    put(ws, 'J', 21, 'ГОД', fill=GREY, font=HDR, border=True)
    for col, t in (('K', 'Выручка'), ('L', 'Прибыль'), ('M', 'Касса на конец')):
        put(ws, col, 21, t, fill=GREY, font=HDR, border=True)
    for rr, (year, (r1, r2)) in enumerate(zip((2026, 2027, 2028, 2029), (Y26, Y27, Y28, Y29))):
        row = 22 + rr
        put(ws, 'J', row, year, fmt='0', border=True, font=HDR)
        put(ws, 'K', row, SUMF(Rng('Q', r1, 'Q', r2)), fmt=MONEY, border=True)
        put(ws, 'L', row, SUMF(Rng('X', r1, 'X', r2)), fmt=MONEY, border=True)
        put(ws, 'M', row, Ref('Y', r2) + Num(0), fmt=MONEY, border=True)
    ws.column_dimensions['J'].width = 36
    ws.column_dimensions['K'].width = 13
    ws.column_dimensions['L'].width = 13
    ws.column_dimensions['M'].width = 13

    return ws


# ============================ 5. СВОДКА ============================

def build_svodka(wb, ref_values):
    sv = wb.create_sheet('Сводка', 0)
    sv.column_dimensions['B'].width = 44
    for c in ('C', 'D', 'E', 'F'):
        sv.column_dimensions[c].width = 19

    put(sv, 'B', 2, 'Финмодель новой точки — лестница сценариев (движок = тренажёр v2)', font=H1)
    put(sv, 'B', 3, 'авг-2026 → дек-2029 · УСН 6% включён · каждый лист — пресет тренажёра, '
                    'параметры можно крутить прямо на листах', font=SMALL)

    names = [p[0] for p in PRESETS]
    put(sv, 'B', 5, 'Сценарий', fill=GREY, font=HDR, border=True)
    for j, nm in enumerate(names):
        put(sv, chr(ord('C') + j), 5, nm, fill=GREY, font=HDR, border=True)

    rows = [
        ('Рычаги пресета', None, None),
        ('Первый месяц в плюсе', 4, DATEF),
        ('Первые 500 тыс./мес', 5, DATEF),
        ('Дно кассы', 6, MONEY),
        ('…когда дно', 7, DATEF),
        ('Не хватает денег (сверх финансирования)', 8, MONEY),
        ('ВСЕГО в проект (финансирование + дыра)', 9, MONEY),
        ('Клиентов к дек-2027', 10, '0'),
        ('Клиентов к дек-2028', 11, '0'),
        ('Клиентов к дек-2029', 12, '0'),
        ('Прибыль за 2028', 13, MONEY),
        ('Прибыль за 2029', 14, MONEY),
        ('КАССА на дек-2029', 15, MONEY),
        ('Точка безубыточности, клиентов', 17, '0'),
    ]
    r = 6
    for label, krow, fmt in rows:
        bold = label.startswith(('КАССА', 'ВСЕГО'))
        put(sv, 'B', r, label, font=HDR if bold else None, border=True)
        for j, (nm, _, desc) in enumerate(PRESETS):
            col = chr(ord('C') + j)
            if krow is None:
                c = put(sv, col, r, desc, font=SMALL, border=True, wrap_text=True)
            else:
                put(sv, col, r, Ref('K', krow, sheet=nm), fmt=fmt, border=True,
                    font=HDR if bold else None)
        r += 1
    sv.row_dimensions[6].height = 42

    # проверка движка: живая касса дек-29 против числа тренажёра
    put(sv, 'B', r, 'Проверка: касса дек-29 = цифре тренажёра?', font=SMALL, border=True)
    for j, (nm, _, _) in enumerate(PRESETS):
        col = chr(ord('C') + j)
        expected = ref_values[nm]['cash_end']
        put(sv, col, r, IF(Fn('ABS', Ref('K', 15, sheet=nm) - Num(round(expected, 2))).lt(5),
                           '✓ как в тренажёре', '≠ (параметры менялись)'), font=SMALL, border=True)
    r += 2

    put(sv, 'B', r, 'ГЛАВНЫЕ ВЫВОДЫ (проверены прогонами — не меняются от рычагов)', fill=GREY, font=HDR)
    r += 1
    for t in [
        '1. Месяц жизни клиента стоит ~3 млн в год: каждый +1 месяц между 3 и 6 ≈ +3 млн/год прибыли '
        'и ~7 млн к кассе-2029. После 6 мес зал упирается в потолок 250.',
        '2. 500 тыс./мес на стартовом прайсе не существуют: даже полный зал даёт ~+320 тыс. '
        '«Пятисотка» = индексации прайса + ~230 клиентов (реально с 2028-го при форсаже).',
        '3. Форсаж окупается за ~2 месяца (клиент ~10 тыс., маржа ~29 тыс.) — но только при жизни ≥ 4,5–5 мес. '
        'Правило: жми, пока зал < 220 и продления подтверждаются; CPL > 1 500 или объём не набирается → откат.',
        '4. Тарифы решают потолок прибыли: все на месячных (5 200 ₽/мес) — полный зал работает в ноль; '
        'дефолт «12 занятий / 2 мес» (~6 150) + единые/индив — зал приносит 2–5 млн в год.',
    ]:
        c = put(sv, 'B', r, t, wrap_text=True)
        sv.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        sv.row_dimensions[r].height = 30
        r += 1
    r += 1

    put(sv, 'B', r, 'ВВОДНЫЕ (июль-2026)', fill=GREY, font=HDR)
    r += 1
    for t in [
        'Капекс 3,25 млн уже больше плана 3,0 — касса в минусе до открытия · органика без программы '
        '≈ 5 покупателей/мес (факт старой точки) · переноса базы нет · продления: ориентир 3 оплаты.',
        'Проверено фактами: CPL 780 ₽ · потолок таргета ~130 лидов/мес. '
        'Ставки, не факты: 220 лидов в форсаже · рефералка 6% · 60% двухмесячных абонементов.',
        'Как пользоваться: жёлтые ячейки на листах — главные рычаги, синие — остальные параметры; '
        'меняешь число — весь лист и сводка пересчитываются. Excel-версия тренажёра: '
        'https://claude.ai/code/artifact/23a6d766-775e-47d0-8847-54a478573d02',
    ]:
        c = put(sv, 'B', r, t, wrap_text=True, font=SMALL)
        sv.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        sv.row_dimensions[r].height = 28
        r += 1
    return sv


# ============================ 6. СВЕРКА ФОРМУЛ С ДВИЖКОМ ============================

def verify_sheets(ref):
    """Вычисляем формулы книги мини-вычислителем и сравниваем с портом движка."""
    worst = 0.0
    for nm, s, _ in PRESETS:
        k = ref[nm]
        r = k['r']
        for i in range(N):
            row = TR0 + i
            for col, expect in (('N', r['A'][i]), ('Q', r['rev'][i]), ('X', r['pnl'][i]), ('Y', r['cash'][i])):
                got = REG.value(nm, col, row)
                d = abs(got - expect)
                worst = max(worst, d)
                assert d < 0.05, f'{nm}!{col}{row}: формула {got} != движок {expect}'
        # KPI
        def kv(row):
            return REG.value(nm, 'K', row)
        exp_fp = MDATES[k['fp']] if k['fp'] is not None else 'нет до 2030'
        exp_f5 = MDATES[k['f5']] if k['f5'] is not None else '—'
        assert kv(4) == exp_fp, f'{nm}: первый плюс {kv(4)} != {exp_fp}'
        assert kv(5) == exp_f5, f'{nm}: 500к {kv(5)} != {exp_f5}'
        assert abs(kv(6) - k['dno']) < 0.05
        assert kv(7) == MDATES[k['dno_i']]
        assert abs(kv(8) - k['shortfall']) < 0.05
        assert abs(kv(9) - k['total']) < 0.05
        assert abs(kv(10) - k['cli27']) < 0.01 and abs(kv(11) - k['cli28']) < 0.01 and abs(kv(12) - k['cli29']) < 0.01
        assert abs(kv(13) - k['y28']) < 0.05 and abs(kv(14) - k['y29']) < 0.05
        assert abs(kv(15) - k['cash_end']) < 0.05
        if not k['dipped']:
            exp16 = 'не проседает'
        elif k['recov'] is None:
            exp16 = 'после 2029'
        else:
            exp16 = MDATES[k['recov']]
        assert kv(16) == exp16, f'{nm}: восстановление {kv(16)} != {exp16}'
        be_cell = kv(17)
        if k['be'] is None:
            assert be_cell == 'недостиж.'
        else:
            assert be_cell == k['be'], f'{nm}: ТБ {be_cell} != {k["be"]}'
        assert abs(REG.value(nm, 'M', 8) - k['be_raw']) < 1e-3
        # годовые итоги
        for rr, (r1, r2) in enumerate((Y26, Y27, Y28, Y29)):
            row = 22 + rr
            assert abs(REG.value(nm, 'K', row) - sum(r['rev'][r1 - TR0:r2 - TR0 + 1])) < 0.05
            assert abs(REG.value(nm, 'L', row) - sum(r['pnl'][r1 - TR0:r2 - TR0 + 1])) < 0.05
            assert abs(REG.value(nm, 'M', row) - r['cash'][r2 - TR0]) < 0.05
    # сводка: проверка движка должна гореть ✓
    for j, (nm, _, _) in enumerate(PRESETS):
        col = chr(ord('C') + j)
        v = REG.value('Сводка', col, 20)
        assert v == '✓ как в тренажёре', f'Сводка {nm}: {v}'
    return worst


# ============================ 7. СБОРКА ============================

def main():
    print('1) Сверка порта движка с цифрами тренажёра:')
    check_port()

    ref = {nm: kpis(s) for nm, s, _ in PRESETS}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for nm, s, desc in PRESETS:
        build_preset_sheet(wb, nm, s, desc)
    build_svodka(wb, ref)

    print('2) Сверка формул Excel с движком (мини-вычислитель):')
    worst = verify_sheets(ref)
    print(f'   все 4 листа × 41 месяц × (клиенты, выручка, результат, касса) + KPI: ✓, '
          f'макс. расхождение {worst:.4f} ₽')

    wb.save(OUT)
    print('3) Сохранено:', OUT)


if __name__ == '__main__':
    main()
