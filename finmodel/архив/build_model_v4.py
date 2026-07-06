# -*- coding: utf-8 -*-
"""Финмодель v2 — Этап 2: листы «Допущения» + «Воронка» + «Читайка».
Все расчёты — живыми формулами Excel, ручной ввод только на «Допущениях».
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

YELLOW = PatternFill("solid", fgColor="FFF2CC")   # ждём данные / заглушка
BLUE   = PatternFill("solid", fgColor="DDEBF7")   # ввод пользователя
GREY   = PatternFill("solid", fgColor="F2F2F2")
HDR    = Font(bold=True)
H1     = Font(bold=True, size=14)
MONEY  = '#,##0'
PCT    = '0%'
thin   = Side(style='thin', color='CCCCCC')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()

# ============================= ЛИСТ «Допущения» =============================
ws = wb.active
ws.title = "Допущения"
ws.column_dimensions['B'].width = 46
for c in ('C',): ws.column_dimensions[c].width = 10
for c in ('D','E','F','G'): ws.column_dimensions[c].width = 12
ws.column_dimensions['H'].width = 68

ws['B2'] = 'Активный сценарий (1=Пессимист, 2=База, 3=Целевой):'
ws['B2'].font = HDR
ws['D2'] = 2
ws['D2'].fill = BLUE
ws['D2'].font = Font(bold=True, size=13)
ws['H2'] = 'Меняешь одну цифру — пересчитывается вся модель. 3 = «к чему стремиться».'

hdr = ['Параметр','ед.','АКТИВНОЕ','Пессимист','База','Целевой','Комментарий / источник']
for i,h in enumerate(hdr):
    cell = ws.cell(row=4, column=2+i, value=h)
    cell.font = HDR
    cell.fill = GREY
    cell.border = BORDER

# (имя, ед., песс, база, цель, коммент, формат, жёлтый?)
params = [
    ('— ВОРОНКА —', None, None, None, None, '', None, False),
    ('CPL — цена сырого лида (таргет)', '₽', 1000, 800, 1200, 'ФАКТ из кабинета ВК (окт-25→май-26, 403 заявки): средний 780 ₽, коридор 700–1100. Целевой дороже: +30% CPL на каждое удвоение объёма.', MONEY, False),
    ('Конверсия платного лида в покупку', '%', 0.10, 0.13, 0.15, 'Факт CRM: VK 13% (n=1039). >15% на чисто платном не бывает.', PCT, False),
    ('Конверсия органического лида в покупку', '%', 0.20, 0.25, 0.25, 'Факт CRM: сайт/органика 25–27% (n=400).', PCT, False),
    ('Потолок платных лидов', 'шт/мес', 70, 100, 200, 'ФАКТ: 30–50/мес на бюджете 30–45т; пик 127 (янв, бюджет 73т, без роста CPL). Выше 130 — экстраполяция = ГЛАВНЫЙ РИСК модели.', MONEY, False),
    ('Органические лиды на старте (сарафан/карты/сайт)', 'шт/мес', 10, 20, 30, 'ЗАГЛУШКА — ждём факт действующего центра.', MONEY, True),
    ('Рост органики', '%/мес', 0.05, 0.08, 0.12, 'ЗАГЛУШКА — реф.программа «привёл друга» (бот) должна разгонять.', PCT, True),
    ('Доля лидов, дошедших до пробного', '%', 0.35, 0.50, 0.55, 'Ваша оценка: ~50%. Для выручки пробных (900 ₽).', PCT, False),
    ('Коэфф. «зачислен → оплатил»', '', 0.93, 0.93, 0.93, 'Факт CRM: 7% зачисленных не платят.', '0.00', False),
    ('— ПРОДУКТ И ДЕНЬГИ —', None, None, None, None, '', None, False),
    ('Средний платёж (один абонемент)', '₽', 9000, 9800, 10500, 'По новому прайсу, микс 70% одно напр. / 20% единый / 10% индив.', MONEY, True),
    ('Средняя длительность абонемента', 'мес', 1.7, 1.7, 1.7, 'Микс 1-мес и 2-мес абонементов.', '0.0', False),
    ('Покупок на клиента за жизнь', 'шт', 1.8, 2.0, 3.0, 'Факт CRM: ~2 (продление 40–49% до 2-й, ~20% до 3-й). Целевой 3 = когорты 2026 (82–94%) — проверить у администратора!', '0.0', False),
    ('Цена пробного', '₽', 900, 900, 900, 'Некупившие платят 900; у купивших зачтено в абонемент.', MONEY, False),
    ('Вместимость: действующих абонементов макс', 'шт', 250, 250, 250, 'Слово собственницы: потолок загрузки 250.', MONEY, False),
    ('Стартовая база клиентов (перенос из действующего центра)', 'чел', 0, 0, 0, 'ОТКРЫТЫЙ ВОПРОС: старт с нуля или с переносом?', MONEY, True),
    ('— РАСХОДЫ (разбивка ФОТ — слова собственницы, июль-2026) —', None, None, None, None, '', None, False),
    ('Прочие постоянные расходы', '₽/мес', 435000, 435000, 435000, 'Из исходника: аренда 360 + РКО 10 + 1С 15 + канц 10 + мотив 30 + связь 10.', MONEY, False),
    ('ФОТ фикс-ядро (маркетинг 130 + админы 120 + руковод. 100)', '₽/мес', 350000, 350000, 350000, 'Платится независимо от числа клиентов — это якорь точки безубыточности.', MONEY, False),
    ('Ставка тренера за группу (даже пустую)', '₽', 500, 500, 500, 'ФАКТ (июль-2026): 500 ₽ за занятие группы, платится даже если никто не пришёл.', MONEY, False),
    ('Ставка тренера за ребёнка', '₽', 100, 100, 100, 'ФАКТ: +100 ₽ за каждого пришедшего ребёнка (посещений ≈ 4,33/мес на клиента).', MONEY, False),
    ('Мин. сетка групп в неделю (стартовое расписание)', 'шт', 24, 30, 30, 'Пустые группы тоже стоят 500 ₽ → floor тренерского ФОТ ≈ 52–65 тыс/мес.', MONEY, True),
    ('Средняя наполняемость группы', 'чел', 3.5, 4.0, 4.5, 'Макс 6 пар; при полном зале тренеры ≈ 244 тыс/мес — бюджет собственницы (250) сходится.', '0.0', False),
    ('— ДЕНЬГИ СОБСТВЕННИЦЫ —', None, None, None, None, '', None, False),
    ('План финансирования всего', '₽', 3000000, 3000000, 3000000, 'Слова собственницы: «рассчитывала на 3 млн» (ремонт + оборудование + первый маркетинг).', MONEY, False),
    ('Уже потрачено до открытия (факт на 03.07.2026)', '₽', 1551733, 1551733, 1551733, 'По списку оплат: ремонт 719,3к · оборудование 303,2к · ИТ/маркетинг 275к · дизайн-проекты 252к · прочее 2,2к.', MONEY, False),
    ('Достройка до открытия (остаток ремонта и оборудования)', '₽', 1700000, 1700000, 1700000, 'ФАКТ (июль-2026): ≈1,7 млн, падает в август. Вместе с потраченными 1,55 млн капекс = 3,25 млн > плана 3,0 млн.', MONEY, False),
]
r = 5
PR = {}  # имя -> строка
for name, unit, p, b, t, comm, fmt, yell in params:
    ws.cell(row=r, column=2, value=name)
    if unit is None:  # заголовок секции
        ws.cell(row=r, column=2).font = HDR
        ws.cell(row=r, column=2).fill = GREY
        r += 1
        continue
    ws.cell(row=r, column=3, value=unit)
    ws.cell(row=r, column=4, value=f'=INDEX(E{r}:G{r},$D$2)')
    for col, v in ((5,p),(6,b),(7,t)):
        ws.cell(row=r, column=col, value=v)
    ws.cell(row=r, column=8, value=comm)
    for col in range(4,8):
        c = ws.cell(row=r, column=col)
        if fmt: c.number_format = fmt
        c.border = BORDER
        if col >= 5: c.fill = BLUE
    ws.cell(row=r, column=4).font = HDR
    if yell:
        for col in range(2,9): ws.cell(row=r, column=col).fill = YELLOW
    PR[name] = r
    r += 1

# производные
r += 1
ws.cell(row=r, column=2, value='— РАСЧЁТНЫЕ (не трогать) —').font = HDR
ws.cell(row=r, column=2).fill = GREY
r += 1
life_row = r
ws.cell(row=r, column=2, value='Жизнь клиента (= покупок × длительность)')
ws.cell(row=r, column=3, value='мес')
ws.cell(row=r, column=4, value=f"=D{PR['Покупок на клиента за жизнь']}*D{PR['Средняя длительность абонемента']}")
ws.cell(row=r, column=4).number_format = '0.0'
ws.cell(row=r, column=4).font = HDR
r += 1
nominal_row = r
ws.cell(row=r, column=2, value='Деньги с действующего клиента в месяц (= платёж / длительность)')
ws.cell(row=r, column=3, value='₽/мес')
ws.cell(row=r, column=4, value=f"=D{PR['Средний платёж (один абонемент)']}/D{PR['Средняя длительность абонемента']}")
ws.cell(row=r, column=4).number_format = MONEY
ws.cell(row=r, column=4).font = HDR
r += 1
budget_row = r
ws.cell(row=r, column=2, value='Бюджет таргета (= потолок лидов × CPL)')
ws.cell(row=r, column=3, value='₽/мес')
ws.cell(row=r, column=4, value=f"=D{PR['Потолок платных лидов']}*D{PR['CPL — цена сырого лида (таргет)']}")
ws.cell(row=r, column=4).number_format = MONEY
ws.cell(row=r, column=4).font = HDR
r += 1
cash0_row = r
ws.cell(row=r, column=2, value='Касса на старте (= план 3 млн − уже потрачено)')
ws.cell(row=r, column=3, value='₽')
ws.cell(row=r, column=4, value=f"=D{PR['План финансирования всего']}-D{PR['Уже потрачено до открытия (факт на 03.07.2026)']}")
ws.cell(row=r, column=4).number_format = MONEY
ws.cell(row=r, column=4).font = HDR

# сезонность
r += 2
ws.cell(row=r, column=2, value='— СЕЗОННОСТЬ (факт вашей CRM; можно править) —').font = HDR
ws.cell(row=r, column=2).fill = GREY
r += 1
for i,h in enumerate(['Месяц','Коэфф. притока лидов','Коэфф. конверсии в покупку']):
    c = ws.cell(row=r, column=2+i, value=h); c.font = HDR; c.fill = GREY
r += 1
season_start = r
season = [  # (мес, приток, конверсия)
    ('январь', 1.0, 0.6), ('февраль', 1.4, 1.0), ('март', 1.1, 1.0), ('апрель', 1.0, 1.0),
    ('май', 1.0, 1.0), ('июнь', 0.6, 0.9), ('июль', 0.5, 0.9), ('август', 0.7, 1.0),
    ('сентябрь', 1.2, 1.0), ('октябрь', 1.1, 1.0), ('ноябрь', 1.0, 1.0), ('декабрь', 0.8, 0.8),
]
for m, inf, conv in season:
    ws.cell(row=r, column=2, value=m).border = BORDER
    ws.cell(row=r, column=3, value=inf).border = BORDER
    ws.cell(row=r, column=4, value=conv).border = BORDER
    ws.cell(row=r, column=3).fill = BLUE
    ws.cell(row=r, column=4).fill = BLUE
    r += 1
ws.cell(row=r, column=2, value='янв: заявки есть, покупают хуже (×0,6); фев: пик (факт ×1,8 — тут консервативно 1,4); лето: приток ×0,5–0,6.')

D = lambda name: f"Допущения!$D${PR[name]}"
LIFE   = f"Допущения!$D${life_row}"
NOMIN  = f"Допущения!$D${nominal_row}"
BUDG   = f"Допущения!$D${budget_row}"
CASH0  = f"Допущения!$D${cash0_row}"
SEAS_I = f"Допущения!$C${season_start}:$C${season_start+11}"
SEAS_C = f"Допущения!$D${season_start}:$D${season_start+11}"

# ============================= ЛИСТ «Воронка» =============================
wf = wb.create_sheet("Воронка")
wf.column_dimensions['B'].width = 44
N = 17  # авг-2026 .. дек-2027
first_col, last_col = 3, 3+N-1           # C..S
y2026 = (3, 7)   # C..G (авг–дек 2026)
y2027 = (8, 19)  # H..S
COLS = [get_column_letter(c) for c in range(first_col, last_col+1)]
T26, T27 = get_column_letter(last_col+1), get_column_letter(last_col+2)

wf['B2'] = 'Воронка и деньги (всё считается с листа «Допущения»)'
wf['B2'].font = H1
# даты
months = []
yy, mm = 2026, 8
for _ in range(N):
    months.append(datetime(yy, mm, 1))
    mm += 1
    if mm == 13: mm, yy = 1, yy+1
for i,d in enumerate(months):
    c = wf.cell(row=3, column=first_col+i, value=d)
    c.number_format = 'MMM YY'; c.font = HDR; c.fill = GREY; c.border = BORDER
wf.cell(row=3, column=last_col+1, value='2026').font = HDR
wf.cell(row=3, column=last_col+2, value='2027').font = HDR
wf.cell(row=3, column=2, value='Месяц').font = HDR

rows_def = []  # (row, label, formula_fn(col_letter, idx), fmt, bold)
def add_row(rr, label, f, fmt=MONEY, bold=False, year_sum=True):
    wf.cell(row=rr, column=2, value=label).font = Font(bold=bold)
    for i, cl in enumerate(COLS):
        c = wf.cell(row=rr, column=first_col+i, value=f(cl, i))
        c.number_format = fmt; c.border = BORDER
        if bold: c.font = HDR
    if year_sum:
        a = wf.cell(row=rr, column=last_col+1,
                    value=f"=SUM({COLS[0]}{rr}:{COLS[4]}{rr})")
        b = wf.cell(row=rr, column=last_col+2,
                    value=f"=SUM({COLS[5]}{rr}:{COLS[-1]}{rr})")
        for c in (a,b):
            c.number_format = fmt
            c.font = HDR

R = {}
r = 5
R['si'] = r; add_row(r, 'Сезонный коэфф. притока',    lambda cl,i: f"=INDEX({SEAS_I},MONTH({cl}$3))", '0.0', year_sum=False); r+=1
R['sc'] = r; add_row(r, 'Сезонный коэфф. конверсии',  lambda cl,i: f"=INDEX({SEAS_C},MONTH({cl}$3))", '0.0', year_sum=False); r+=1
r += 1
R['pl'] = r; add_row(r, 'Платные лиды (таргет, с потолком канала)',
    lambda cl,i: f"=MIN({D('Потолок платных лидов')},{BUDG}/{D('CPL — цена сырого лида (таргет)')})*{cl}{R['si']}"); r+=1
R['ol'] = r; add_row(r, 'Органические лиды (сарафан/карты/сайт)',
    lambda cl,i: f"={D('Органические лиды на старте (сарафан/карты/сайт)')}*(1+{D('Рост органики')})^{i}*{cl}{R['si']}"); r+=1
R['tl'] = r; add_row(r, 'Лидов всего', lambda cl,i: f"={cl}{R['pl']}+{cl}{R['ol']}", bold=True); r+=1
R['tr'] = r; add_row(r, 'Пробных занятий',
    lambda cl,i: f"=({cl}{R['pl']}+{cl}{R['ol']})*{D('Доля лидов, дошедших до пробного')}*{cl}{R['sc']}"); r+=1
R['nc'] = r; add_row(r, 'НОВЫХ клиентов (купили абонемент)',
    lambda cl,i: (f"=({cl}{R['pl']}*{D('Конверсия платного лида в покупку')}"
                  f"+{cl}{R['ol']}*{D('Конверсия органического лида в покупку')})"
                  f"*{cl}{R['sc']}*{D('Коэфф. «зачислен → оплатил»')}"), bold=True); r+=1
R['ac'] = r
def actives(cl, i):
    prev = f"{D('Стартовая база клиентов (перенос из действующего центра)')}" if i==0 else f"{COLS[i-1]}{R['ac']}"
    return (f"=MIN({D('Вместимость: действующих абонементов макс')},"
            f"{prev}*(1-1/{LIFE})+{cl}{R['nc']})")
add_row(r, 'ДЕЙСТВУЮЩИХ клиентов (абонементы)', actives, bold=True, year_sum=False); r+=1
R['load'] = r; add_row(r, 'Загрузка от потолка 250',
    lambda cl,i: f"={cl}{R['ac']}/{D('Вместимость: действующих абонементов макс')}", PCT, year_sum=False); r+=1
r += 1
R['ra'] = r; add_row(r, 'Выручка: абонементы', lambda cl,i: f"={cl}{R['ac']}*{NOMIN}"); r+=1
R['rt'] = r; add_row(r, 'Выручка: пробные (некупившие)',
    lambda cl,i: f"=MAX(0,{cl}{R['tr']}-{cl}{R['nc']})*{D('Цена пробного')}"); r+=1
R['rev'] = r; add_row(r, 'ИТОГО выручка', lambda cl,i: f"={cl}{R['ra']}+{cl}{R['rt']}", bold=True); r+=1
r += 1
R['mkt'] = r; add_row(r, 'Маркетинг (таргет)', lambda cl,i: f"={BUDG}"); r+=1
R['opx'] = r; add_row(r, 'Прочие постоянные (аренда и т.п.)',
    lambda cl,i: f"={D('Прочие постоянные расходы')}"); r+=1
R['fot'] = r; add_row(r, 'ФОТ фикс-ядро (маркетинг+админы+руковод.)',
    lambda cl,i: f"={D('ФОТ фикс-ядро (маркетинг 130 + админы 120 + руковод. 100)')}"); r+=1
R['tea'] = r; add_row(r, 'ФОТ преподаватели (500/группа + 100/ребёнок)',
    lambda cl,i: (f"=MAX({D('Мин. сетка групп в неделю (стартовое расписание)')},{cl}{R['ac']}/{D('Средняя наполняемость группы')})"
                  f"*4.33*{D('Ставка тренера за группу (даже пустую)')}"
                  f"+{cl}{R['ac']}*4.33*{D('Ставка тренера за ребёнка')}")); r+=1
R['tax'] = r
taxes = {0: 38260, 4: 40000}  # авг-26 (фикс. взносы ИП), дек-26 (как в исходнике)
add_row(r, 'Налоги (ровно как в исходнике!)',
    lambda cl,i: taxes.get(i, 0)); r+=1
wf.cell(row=R['tax'], column=2+18).value = '⚠ 2027 в исходнике не было: УСН/взносы не заложены'
R['pnl'] = r; add_row(r, 'ОПЕРАЦИОННЫЙ РЕЗУЛЬТАТ (месяц)',
    lambda cl,i: f"={cl}{R['rev']}-{cl}{R['mkt']}-{cl}{R['opx']}-{cl}{R['fot']}-{cl}{R['tea']}-{cl}{R['tax']}", bold=True); r+=1
r += 1
R['cap'] = r; add_row(r, 'Достройка до открытия (капекс, август)',
    lambda cl,i: (f"={D('Достройка до открытия (остаток ремонта и оборудования)')}" if i==0 else "=0")); r+=1
R['cash'] = r
def cash(cl, i):
    prev = f"{CASH0}" if i==0 else f"{COLS[i-1]}{R['cash']}"
    return f"={prev}+{cl}{R['pnl']}-{cl}{R['cap']}"
add_row(r, 'КАССА на конец месяца', cash, bold=True, year_sum=False)

# красная подсветка отрицательной кассы
from openpyxl.formatting.rule import CellIsRule
RED = PatternFill("solid", fgColor="F8CBAD")
rng = f"{COLS[0]}{R['cash']}:{COLS[-1]}{R['cash']}"
wf.conditional_formatting.add(rng, CellIsRule(operator='lessThan', formula=['0'], fill=RED))

# ============================= ЛИСТ «Читайка» =============================
wr = wb.create_sheet("Читайка")
wr.column_dimensions['B'].width = 110
txt = [
    ('КАК ПОЛЬЗОВАТЬСЯ', True),
    ('1. Лист «Допущения», ячейка D2: переключи сценарий 1/2/3 — вся модель пересчитается.', False),
    ('2. Синие ячейки — можно править. Жёлтые строки — заглушки, ждём реальные данные.', False),
    ('3. Лист «Воронка»: главные строки — ДЕЙСТВУЮЩИХ клиентов, ИТОГО выручка, ОПЕРАЦИОННЫЙ РЕЗУЛЬТАТ.', False),
    ('', False),
    ('ЧТО ЭТО ЗА ЭТАП', True),
    ('Это ГОТОВАЯ рабочая версия (этапы 1–4): сводка + воронка + деньги. ФОТ разнесён (фикс-ядро 350к + преподаватели по загрузке),', False),
    ('заведены деньги собственницы (план 3 млн, потрачено 1 551 733) и строка «КАССА на конец месяца» —', False),
    ('красная заливка = деньги кончились. Налоги — ровно как в исходнике (2027 пустой — там налогов не было!).', False),
    ('Этап 4: сводка сценариев — сколько всего нужно денег, когда ноль, когда окупаемость.', False),
    ('', False),
    ('ЧЕГО НЕ ХВАТАЕТ (жёлтые места)', True),
    ('Подтверждено (июль-2026): открытие 1–5 августа; достройка ≈1,7 млн; тренеры 500 ₽/группа + 100 ₽/ребёнок.', False),
    ('— Органика: сколько лидов/мес сейчас без рекламы (сарафан/карты/сайт)', False),
    ('— Старт: с нуля или с переносом базы действующего центра', False),
    ('— Таргетолог: максимум лидов/мес и на каком бюджете; администратор: продления 82–94% — правда или формат оплаты', False),
    ('— «Нежданчик» — вводные для прогноза', False),
]
rr = 2
for t, bold in txt:
    c = wr.cell(row=rr, column=2, value=t)
    if bold: c.font = HDR
    rr += 1


# ============================= ЛИСТ «Сводка» =============================
sv = wb.create_sheet("Сводка", 0)   # первым листом
sv.column_dimensions['B'].width = 52
sv.column_dimensions['C'].width = 16
for c in ('D','E','F','G'): sv.column_dimensions[c].width = 17
sv['B2'] = 'СВОДКА — главное одним экраном'
sv['B2'].font = H1

sv['B4'] = 'ТЕКУЩИЙ СЦЕНАРИЙ (живое: пересчитывается при правке «Допущений»)'
sv['B4'].font = HDR; sv['B4'].fill = GREY
live = [
    ('Выбран сценарий (1=Пессимист, 2=База, 3=Целевой)', '=Допущения!$D$2', '0'),
    ('Клиентов к декабрю 2027', f"=Воронка!S{R['ac']}", MONEY),
    ('Выручка за 2027', f"=Воронка!U{R['rev']}", MONEY),
    ('Операционный результат за 2027', f"=Воронка!U{R['pnl']}", MONEY),
    ('Месяцев в минусе (из 17)', f"=COUNTIF(Воронка!C{R['pnl']}:S{R['pnl']},\"<0\")", '0'),
    ('Дно кассы (самая глубокая яма)', f"=MIN(Воронка!C{R['cash']}:S{R['cash']})", MONEY),
    ('НУЖНО ДЕНЕГ сверх плана 3 млн', f"=MAX(0,-MIN(Воронка!C{R['cash']}:S{R['cash']}))", MONEY),
    ('Касса на конец 2027', f"=Воронка!S{R['cash']}", MONEY),
]
rr = 5
for label, f, fmt in live:
    sv.cell(row=rr, column=2, value=label)
    c = sv.cell(row=rr, column=3, value=f); c.number_format = fmt; c.font = HDR; c.border = BORDER
    rr += 1

rr += 1
sv.cell(row=rr, column=2, value='ТРИ СЦЕНАРИЯ РЯДОМ (снимок при допущениях от 03.07.2026 — после правок сверяйся с живым блоком выше)').font = HDR
sv.cell(row=rr, column=2).fill = GREY
rr += 1
head = ['Сценарий','Клиентов дек-27','Результат/мес в конце 27','Первый плюсовой месяц','Нужно сверх 3 млн']
for i,h in enumerate(head):
    c = sv.cell(row=rr, column=2+i, value=h); c.font = HDR; c.border = BORDER; c.fill = GREY
rr += 1
for row in [('Пессимист','27','−751 тыс.','никогда','13,6 млн'),
            ('База (историч. цифры)','75','−474 тыс.','никогда','10,5 млн'),
            ('Целевой (все рычаги)','241','+359 тыс.','март 2027','3,6 млн')]:
    for i,v in enumerate(row):
        c = sv.cell(row=rr, column=2+i, value=v); c.border = BORDER
    rr += 1

rr += 1
sv.cell(row=rr, column=2, value='РЫЧАГИ: «База» + один рычаг за раз (что двигает результат)').font = HDR
sv.cell(row=rr, column=2).fill = GREY
rr += 1
for i,h in enumerate(['Вариант','Клиентов дек-27','Результат/мес','Нужно сверх 3 млн']):
    c = sv.cell(row=rr, column=2+i, value=h); c.font = HDR; c.border = BORDER; c.fill = GREY
rr += 1
for row in [('База как есть','75','−474 тыс.','10,5 млн'),
            ('+ продления как у когорт-2026 (3 покупки)','106','−317 тыс.','8,9 млн'),
            ('+ органика 50 лидов/мес (потолок 100)','103','−297 тыс.','7,9 млн'),
            ('+ конверсия платного 15%','80','−445 тыс.','10,1 млн'),
            ('+ таргет 200 лидов/мес','110','−322 тыс.','8,4 млн'),
            ('ПРОДЛЕНИЯ + ОРГАНИКА (таргет не трогаем)','150','−77 тыс.','5,6 млн'),
            ('Все рычаги вместе','218','+211 тыс.','2,9 млн'),]:
    for i,v in enumerate(row):
        c = sv.cell(row=rr, column=2+i, value=v); c.border = BORDER
    rr += 1

rr += 1
sv.cell(row=rr, column=2, value='ЧТО УТОЧНИТЬ, ЧТОБЫ ЦИФРЫ СТАЛИ ТВЁРДЫМИ (сейчас — заглушки, жёлтое на «Допущениях»)').font = HDR
sv.cell(row=rr, column=2).fill = GREY
rr += 1
for q in ['ПОДТВЕРЖДЕНО (июль-2026): открытие 1–5 августа · достройка ≈1,7 млн · тренеры 500 ₽/группа + 100 ₽/ребёнок',
          '1. Продления 82–94% у когорт-2026 — правда или формат оплаты (администратор) — рычаг №1',
          '2. Органика: сколько лидов/мес бесплатно приходит сейчас — рычаг №2',
          '3. Таргетолог: максимум лидов/мес и на каком бюджете',
          '4. Старт с нуля или с переносом части базы действующего центра',
          '5. Какая CRM куплена за 103 тыс/год',
          '6. Финансирование: капекс уже больше плана — сколько денег реально доступно и когда?']:
    sv.cell(row=rr, column=2, value=q).fill = YELLOW
    rr += 1

rr += 1
for concl in ['ВЫВОДЫ: капекс 3,25 млн (потрачено 1,55 + достройка 1,7) УЖЕ больше плана 3,0 — касса в минусе ещё до открытия.',
              'ТБ ≈ 185–190 клиентов. «База» бездонна (−11,2 млн к дек-2027). Путь в плюс = продления + сарафан + чек; целевой: плюс с марта 2027.',
              'Разговор о деньгах — не «до октября», а сейчас: без ещё ~3,5 млн подтверждённых даже лучший план умирает в сентябре.']:
    sv.cell(row=rr, column=2, value=concl).font = HDR
    rr += 1


out = '/Users/nikolaymacbook/Claude/mama-ryadom-bot/Финмодель v2.xlsx'
wb.save(out)
print('saved:', out)
